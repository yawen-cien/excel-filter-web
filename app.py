import io
import os
import pandas as pd
import streamlit as st
from datetime import datetime
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# --- 網頁頁面設定 ---
st.set_page_config(page_title="Excel 自動化篩選工具", page_icon="📊", layout="centered")

st.title("📊 Excel 自動化篩選工具 Web 版")
st.write("請先輸入篩選條件，再上傳您的 Excel 檔案。系統會自動進行健全性檢核並排版。")

# --- 側邊欄：篩選條件設定 ---
st.sidebar.header("⚙️ 篩選條件設定")
target_column = st.sidebar.text_input("1. 請輸入 Excel 欄位名稱:", value="地區").strip()
target_keyword = st.sidebar.text_input("2. 請輸入篩選關鍵字:", value="北部").strip()

# --- 主畫面：檔案上傳 ---
# 🛠️ 這裡已修正為正確的 st.file_uploader 指令
uploaded_file = st.file_uploader("選擇要篩選的 Excel 檔案", type=["xlsx", "xls"])

if uploaded_file is not None:
    try:
        # 1. 讀取 Excel 資料
        df = pd.read_excel(uploaded_file)
        st.success("📂 檔案上傳成功！")
        
        # 呈現原始資料預覽（前 5 筆）
        with st.expander("🔍 點擊查看原始資料預覽 (前 5 筆)"):
            st.dataframe(df.head())

        # 🛑 檢核：提示無可篩選的相關欄位
        if target_column not in df.columns:
            st.error(f"❌ 錯誤：上傳的 Excel 中找不到『{target_column}』欄位！無法產生結果。")
            st.stop()

        # 2. 執行動態篩選
        condition = df[target_column].astype(str) == target_keyword
        filtered_df = df[condition]

        # 🛑 檢核：若篩選後完全沒有符合的資料
        if filtered_df.empty:
            st.warning(f"⚠️ 警告：找不到『{target_column}』欄位符合『{target_keyword}』的資料！無法產生結果。")
            st.stop()

        # 呈現篩選結果預覽
        st.write(f"✨ 成功篩選出 **{len(filtered_df)}** 筆符合的資料：")
        st.dataframe(filtered_df)

        # 3. 在記憶體中建立 Excel（使用 BytesIO），並進行最適欄寬與靠左對齊優化
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            filtered_df.to_excel(writer, sheet_name="篩選結果", index=False)
            worksheet = writer.sheets["篩選結果"]
            
            # 定義排版對齊樣式
            center_alignment = Alignment(horizontal="center", vertical="center") # 標題置中
            left_alignment = Alignment(horizontal="left", vertical="center")     # 內文置左

            # 安全巡迴每個儲存格
            for col_idx in range(1, worksheet.max_column + 1):
                max_len = 0
                col_letter = get_column_letter(col_idx)

                for row_idx in range(1, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    
                    if row_idx == 1:
                        cell.alignment = center_alignment  # 標題置中
                    else:
                        cell.alignment = left_alignment    # 內容置左
                    
                    val_str = str(cell.value or "")
                    if len(val_str) > max_len:
                        max_len = len(val_str)

                worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # 準備下載檔案的數據
        excel_data = output.getvalue()

        # 生成帶有時間戳記的獨立檔名
        current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
        download_filename = f"篩選結果_{target_column}_{target_keyword}_{current_time}.xlsx"

        # 4. 提供網頁下載按鈕
        st.download_button(
            label="📥 點擊下載篩選與排版優化後的 Excel",
            data=excel_data,
            file_name=download_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"處理過程中遇到錯誤：{str(e)}")